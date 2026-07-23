"""
Atualizar dashboard RTA a partir das planilhas Excel de origem
=================================================================

O que este script faz:
  1. Le as DUAS planilhas de origem: Unidade e Dados (cada uma pode
     estar na pasta indicada, ou voce pode apontar direto pro arquivo).
  2. Converte os dados para o formato usado pelo dashboard HTML.
  3. Substitui APENAS os dois blocos de dados (unidadeRaw / dadosRaw)
     dentro do arquivo HTML, mantendo todo o resto do arquivo
     (layout, CSS, graficos, filtros, etc.) exatamente como esta.

Como usar
---------
1. Instale as dependencias (uma unica vez):
       pip install openpyxl

2. Rode o script apontando para as duas planilhas e o HTML:
       python atualizar_dashboard_rta.py --unidade "CAMINHO\\Unidade" --dados "CAMINHO\\Dados" --html "CAMINHO\\DO\\DASHBOARD.html"

   Exemplo com os seus caminhos reais (pode apontar pra pasta OU pro
   arquivo .xlsx especifico dentro dela):
       python atualizar_dashboard_rta.py ^
           --unidade "C:\\Users\\raulribeiro\\OneDrive - CLEALCO AÇÚCAR E ÁLCOOL S.A\\Compartilhados\\Base de dados RTA\\Dados Excel\\Unidade" ^
           --dados "C:\\Users\\raulribeiro\\OneDrive - CLEALCO AÇÚCAR E ÁLCOOL S.A\\Compartilhados\\Base de dados RTA\\Dados Excel\\Dados" ^
           --html "C:\\Users\\raulribeiro\\Documents\\GitHub\\App RTA\\index.html"

3. Rodando sem nenhum argumento (ex: botao "Run" do VS Code), o script
   pergunta os caminhos na primeira vez e salva num config_dashboard_rta.json
   ao lado do script -- nas proximas vezes, e so rodar de novo.

4. O HTML sera atualizado no proprio arquivo (por padrao). Se quiser
   gerar um arquivo novo em vez de sobrescrever, use --output.

Requisitos das planilhas
--------------------------
Planilha "Unidade" precisa ter as colunas (em qualquer ordem, pelo nome
do cabecalho, em qualquer aba): SERIAL, Emitente, Gestor aprovador,
Data, Turno, Classificacao, Indicador, e a coluna de status de
aprovacao (ParecerGestor ou Aprovou).

Planilha "Dados" precisa ter as colunas: RTA_SERIAL, ACAO, Responsavel,
SITUACAO.

Modo legado (uma unica planilha com as duas abas "Unidade" e "Dados"):
ainda funciona, use --xlsx "caminho\\Base Juntada.xlsx" no lugar de
--unidade/--dados.
"""

import argparse
import datetime
import json
import os
import re
import stat
import subprocess
import sys
from pathlib import Path

try:
    import openpyxl
except ImportError:
    sys.exit(
        "Faltou instalar a biblioteca 'openpyxl'.\n"
        "Rode: pip install openpyxl"
    )


UNIDADE_COLS = ["SERIAL", "Emitente", "Gestor aprovador", "Data", "Turno",
                "Classificacao", "Indicador"]
# A coluna de status de aprovacao pode ter nomes diferentes dependendo da
# planilha. Tenta nesta ordem de prioridade (o primeiro nome que existir
# na planilha vence).
STATUS_HEADER_CANDIDATOS = ["ParecerGestor", "Aprovou"]
DADOS_COLS = ["RTA_SERIAL", "ACAO", "Responsavel", "SITUACAO"]


def clean(v):
    """Converte valores do Excel para tipos simples (str/None), formata datas."""
    if v is None:
        return None
    if isinstance(v, (datetime.datetime, datetime.date)):
        return v.strftime("%Y-%m-%d")
    if isinstance(v, str):
        v = v.strip()
        return v if v != "" else None
    return v


def find_header_map(header_row, expected_cols, sheet_name):
    """Mapeia nome da coluna -> indice, aceitando nomes com espacos/maiusculas variando.

    Se o mesmo cabecalho aparecer em mais de uma coluna (ex: 'Aprovou' repetido),
    mantem a PRIMEIRA ocorrencia (coluna mais a esquerda) e avisa sobre a duplicidade,
    em vez de deixar a ultima sobrescrever silenciosamente.
    """
    normalized = {}
    all_occurrences = {}
    for idx, cell in enumerate(header_row):
        if cell is None:
            continue
        key = str(cell).strip().lower()
        all_occurrences.setdefault(key, []).append(idx)
        if key not in normalized:
            normalized[key] = idx  # mantem a primeira ocorrencia

    mapping = {}
    missing = []
    for col in expected_cols:
        key = col.strip().lower()
        if key in normalized:
            mapping[col] = normalized[key]
        else:
            missing.append(col)

    if missing:
        raise ValueError(
            f"Na aba '{sheet_name}' nao encontrei a(s) coluna(s): {', '.join(missing)}.\n"
            f"Cabecalhos encontrados: {[c for c in header_row]}"
        )

    # Aviso de transparencia: mostra qual coluna (letra) foi usada para cada campo,
    # e alerta se havia cabecalhos duplicados na planilha.
    print(f"  Colunas usadas na aba '{sheet_name}':")
    for col in expected_cols:
        idx = mapping[col]
        letra = openpyxl.utils.get_column_letter(idx + 1)
        print(f"    {col:20s} -> coluna {letra}")
    for key, idxs in all_occurrences.items():
        if len(idxs) > 1:
            letras = [openpyxl.utils.get_column_letter(i + 1) for i in idxs]
            primeira = openpyxl.utils.get_column_letter(idxs[0] + 1)
            print(f"  AVISO: cabecalho '{key}' aparece em mais de uma coluna ({', '.join(letras)}). "
                  f"Usando a coluna {primeira} (primeira ocorrencia).")

    return mapping


def find_status_column(header_row, candidatos):
    """Procura a coluna de status de aprovacao entre varios nomes possiveis
    de cabecalho, na ordem de prioridade de 'candidatos'. Retorna (nome_usado, indice)."""
    normalized = {}
    for idx, cell in enumerate(header_row):
        if cell is None:
            continue
        key = str(cell).strip().lower()
        if key not in normalized:
            normalized[key] = idx

    for nome in candidatos:
        key = nome.strip().lower()
        if key in normalized:
            return nome, normalized[key]

    raise ValueError(
        f"Nao encontrei nenhuma coluna de status de aprovacao. "
        f"Procurei por: {', '.join(candidatos)}.\n"
        f"Cabecalhos encontrados: {[c for c in header_row]}"
    )


def extract_unidade(ws):
    header = [c.value for c in ws[1]]
    colmap = find_header_map(header, UNIDADE_COLS, "Unidade")

    status_nome, status_idx = find_status_column(header, STATUS_HEADER_CANDIDATOS)
    status_letra = openpyxl.utils.get_column_letter(status_idx + 1)
    print(f"    {'Status aprovacao':20s} -> coluna {status_letra}  (cabecalho: '{status_nome}')")

    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        serial = row[colmap["SERIAL"]]
        if not serial:
            continue
        # Celula de status vazia = trata como "Nao Aprovado" (a base so grava
        # "Aprovado" ou "Solicitar Ajuste"; quando fica em branco, e um RTA
        # que ainda nao foi aprovado).
        aprovou = clean(row[status_idx]) or "Não Aprovado"
        rows.append({
            "SERIAL": clean(serial),
            "Emitente": clean(row[colmap["Emitente"]]),
            "Gestor aprovador": clean(row[colmap["Gestor aprovador"]]),
            "Data": clean(row[colmap["Data"]]),
            "Turno": clean(row[colmap["Turno"]]),
            "Classificacao": clean(row[colmap["Classificacao"]]),
            "Indicador": clean(row[colmap["Indicador"]]),
            "Aprovou": aprovou,

        })
    return rows


def extract_dados(ws):
    header = [c.value for c in ws[1]]
    colmap = find_header_map(header, DADOS_COLS, "Dados")

    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        rta_serial = row[colmap["RTA_SERIAL"]]
        if not rta_serial:
            continue
        rows.append({
            "RTA_SERIAL": clean(rta_serial),
            "ACAO": clean(row[colmap["ACAO"]]),
            "Responsavel": clean(row[colmap["Responsavel"]]),
            "SITUACAO": clean(row[colmap["SITUACAO"]]),
        })
    return rows


def escolher_aba(wb, candidatos_nomes, rotulo):
    """Escolhe a aba certa dentro do workbook: tenta os nomes candidatos
    (case-insensitive); se nenhum bater, usa a primeira aba do arquivo
    (comum quando a planilha vem de uma exportacao tipo 'output')."""
    normalizados = {sn.strip().lower(): sn for sn in wb.sheetnames}
    for nome in candidatos_nomes:
        key = nome.strip().lower()
        if key in normalizados:
            aba_real = normalizados[key]
            print(f"  Aba usada para '{rotulo}': '{aba_real}'")
            return wb[aba_real]
    # fallback: primeira aba do arquivo
    aba_real = wb.sheetnames[0]
    print(f"  Aba usada para '{rotulo}': '{aba_real}' (primeira aba do arquivo)")
    return wb[aba_real]


def resolver_planilha(caminho: Path, rotulo: str) -> Path:
    """Se 'caminho' for uma pasta, procura um unico .xlsx dentro dela
    (ignorando arquivos temporarios de lock do Excel, tipo ~$arquivo.xlsx).
    Se for um arquivo, apenas confirma que existe."""
    if caminho.is_dir():
        xlsxs = sorted(p for p in caminho.glob("*.xlsx") if not p.name.startswith("~$"))
        if len(xlsxs) == 1:
            print(f"AVISO: '{caminho}' e uma pasta. Encontrei um unico .xlsx: {xlsxs[0].name}")
            return xlsxs[0].resolve()
        elif len(xlsxs) > 1:
            lista = "\n".join(f"  - {p.name}" for p in xlsxs)
            sys.exit(
                f"ERRO: a pasta de '{rotulo}' ({caminho}) tem mais de um .xlsx:\n{lista}\n\n"
                "Informe o caminho completo do arquivo certo, incluindo o nome."
            )
        else:
            sys.exit(f"ERRO: a pasta de '{rotulo}' ({caminho}) nao tem nenhum .xlsx dentro dela.")
    elif not caminho.exists():
        sys.exit(f"ERRO: planilha de '{rotulo}' nao encontrada em: {caminho}")
    return caminho


def abrir_planilha(caminho: Path, rotulo: str):
    print(f"Lendo planilha ({rotulo}): {caminho}")
    try:
        return openpyxl.load_workbook(caminho, data_only=True)
    except PermissionError:
        sys.exit(
            f"ERRO: nao consegui abrir a planilha de '{rotulo}':\n  {caminho}\n\n"
            "Isso quase sempre significa que o arquivo esta aberto no Excel\n"
            "(ou outro programa) no seu computador -- o Excel trava o arquivo\n"
            "para leitura enquanto ele estiver aberto.\n\n"
            "Feche a planilha no Excel e rode o script de novo."
        )
    except Exception as e:
        sys.exit(f"ERRO ao abrir a planilha de '{rotulo}': {e}")


UNIDADE_SHEET_CANDIDATOS = ["Unidade", "output"]
DADOS_SHEET_CANDIDATOS = ["Dados", "output"]


def update_html(html_path: Path, unidade_json: str, dados_json: str, output_path: Path):
    content = html_path.read_text(encoding="utf-8")

    new_content, n1 = re.subn(
        r"const unidadeRaw = \[.*?\];",
        "const unidadeRaw = " + unidade_json + ";",
        content, count=1, flags=re.S,
    )
    new_content, n2 = re.subn(
        r"const dadosRaw = \[.*?\];",
        "const dadosRaw = " + dados_json + ";",
        new_content, count=1, flags=re.S,
    )

    if n1 == 0 or n2 == 0:
        raise ValueError(
            "Nao encontrei 'const unidadeRaw = [...]' e/ou 'const dadosRaw = [...]' "
            "no HTML informado. Confirme se e o arquivo certo do dashboard."
        )

    output_path.write_text(new_content, encoding="utf-8")


CONFIG_FILE = Path(__file__).resolve().parent / "config_dashboard_rta.json"


def carregar_config():
    if CONFIG_FILE.exists():
        try:
            return json.loads(CONFIG_FILE.read_text(encoding="utf-8"))
        except Exception:
            return {}
    return {}


def salvar_config(unidade, dados, html, output, xlsx=None):
    cfg = {"html": str(html)}
    if xlsx:
        cfg["xlsx"] = str(xlsx)
    else:
        cfg["unidade"] = str(unidade)
        cfg["dados"] = str(dados)
    if output:
        cfg["output"] = str(output)
    CONFIG_FILE.write_text(json.dumps(cfg, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"Caminhos salvos em: {CONFIG_FILE}")
    print("(da proxima vez, basta rodar o script sem digitar nada -- ele reusa esses caminhos)")


def perguntar_caminho(pergunta, valor_atual):
    sufixo = f" [ENTER = {valor_atual}]" if valor_atual else ""
    resposta = input(f"{pergunta}{sufixo}: ").strip().strip('"')
    return resposta if resposta else valor_atual


def encontrar_repo_git(caminho: Path):
    """Sobe a partir de 'caminho' procurando uma pasta .git (raiz do repositorio).
    Retorna o Path da raiz do repo, ou None se nao encontrar."""
    atual = caminho if caminho.is_dir() else caminho.parent
    for _ in range(8):
        if (atual / ".git").exists():
            return atual
        if atual.parent == atual:
            break
        atual = atual.parent
    return None


def git_sincronizar(repo_dir: Path):
    """Roda git add / commit / push na pasta do repositorio. Nunca trava o
    script com erro fatal -- se algo der errado, so avisa e segue."""

    def rodar(cmd):
        try:
            return subprocess.run(cmd, cwd=repo_dir, capture_output=True, text=True)
        except FileNotFoundError:
            return None

    print("-" * 60)
    print(f"Sincronizando com o GitHub (repositorio: {repo_dir})")

    r = rodar(["git", "add", "."])
    if r is None:
        print("AVISO: nao encontrei o comando 'git'. Instale o Git for Windows")
        print("(git-scm.com) se quiser que o script envie pro GitHub sozinho.")
        return
    if r.returncode != 0:
        print("AVISO: 'git add' falhou:")
        print((r.stdout + r.stderr).strip())
        return

    msg = f"Atualiza dados do dashboard - {datetime.datetime.now().strftime('%d/%m/%Y %H:%M')}"
    r = rodar(["git", "commit", "-m", msg])
    saida = (r.stdout + r.stderr).lower()
    if r.returncode != 0:
        if "nothing to commit" in saida or "nada a submeter" in saida or "nada para submeter" in saida:
            print("Nenhuma mudanca nova para enviar (dados iguais aos ja publicados no GitHub).")
            return
        print("AVISO: 'git commit' falhou:")
        print((r.stdout + r.stderr).strip())
        return
    print(f"Commit criado: \"{msg}\"")

    r = rodar(["git", "push"])
    if r.returncode != 0:
        print("AVISO: 'git push' falhou. Mensagem do git:")
        print((r.stdout + r.stderr).strip())
        print("Voce pode resolver o problema e rodar 'git push' manualmente na pasta do repositorio.")
        return
    print("Enviado para o GitHub com sucesso! (o site publicado sera atualizado em alguns instantes)")


def main():
    parser = argparse.ArgumentParser(description="Atualiza o dashboard RTA (HTML) com dados das planilhas Excel.")
    parser.add_argument("--unidade", required=False, default=None, help="Caminho da planilha (ou pasta) de Unidade")
    parser.add_argument("--dados", required=False, default=None, help="Caminho da planilha (ou pasta) de Dados")
    parser.add_argument("--xlsx", required=False, default=None,
                         help="[Legado] Caminho de uma unica planilha com as abas 'Unidade' e 'Dados'")
    parser.add_argument("--html", required=False, default=None, help="Caminho do arquivo HTML do dashboard a atualizar")
    parser.add_argument("--output", required=False, default=None,
                         help="Caminho de saida (opcional). Se omitido, sobrescreve o proprio --html")
    parser.add_argument("--no-push", action="store_true",
                         help="Nao enviar automaticamente para o GitHub (so atualiza o HTML localmente)")
    args = parser.parse_args()

    config = carregar_config()
    unidade_arg = args.unidade or config.get("unidade")
    dados_arg = args.dados or config.get("dados")
    xlsx_arg = args.xlsx or config.get("xlsx")
    html_arg = args.html or config.get("html")
    output_arg = args.output or config.get("output")

    # Se rodou sem nenhum argumento (ex: botao "Run" do VS Code) e nao ha
    # config salva ainda, pergunta os caminhos interativamente.
    sem_args_cli = not (args.unidade or args.dados or args.xlsx or args.html)
    if sem_args_cli:
        print("Rodando sem parametros de linha de comando -- vou perguntar os caminhos.")
        print("(Dica: voce tambem pode rodar com --unidade \"...\" --dados \"...\" --html \"...\" pelo terminal)")
        print("-" * 60)
        if xlsx_arg and not (unidade_arg or dados_arg):
            # config antiga (modo legado) -- so confirma, nao forca migrar
            xlsx_arg = perguntar_caminho("Caminho da planilha Excel (Base Juntada)", xlsx_arg)
        else:
            unidade_arg = perguntar_caminho("Caminho da planilha/pasta de Unidade", unidade_arg)
            dados_arg = perguntar_caminho("Caminho da planilha/pasta de Dados", dados_arg)
        html_arg = perguntar_caminho("Caminho do HTML do dashboard", html_arg)

        modo_legado_ok = xlsx_arg and not (unidade_arg and dados_arg)
        modo_novo_ok = unidade_arg and dados_arg
        if not html_arg or not (modo_legado_ok or modo_novo_ok):
            sys.exit("ERRO: preciso dos caminhos das planilhas e do HTML para continuar.")
        salvar_config(unidade_arg, dados_arg, html_arg, output_arg, xlsx=xlsx_arg if modo_legado_ok else None)

    usar_legado = bool(xlsx_arg) and not (unidade_arg and dados_arg)

    if not html_arg:
        sys.exit("ERRO: informe --html \"caminho\\do\\dashboard.html\"")
    if not usar_legado and not (unidade_arg and dados_arg):
        sys.exit("ERRO: informe --unidade \"...\" e --dados \"...\" (ou --xlsx \"...\" no modo legado)")

    html_path = Path(html_arg).resolve()
    output_path = Path(output_arg).resolve() if output_arg else html_path

    print("-" * 60)
    print(f"HTML de entrada (caminho completo): {html_path}")
    print(f"HTML de saida (caminho completo): {output_path}")
    print("-" * 60)

    # Se apontou para uma PASTA em vez do arquivo .html, tenta descobrir
    # automaticamente qual .html usar.
    if html_path.is_dir():
        htmls = sorted(html_path.glob("*.html"))
        if len(htmls) == 1:
            print(f"AVISO: '{html_path}' e uma pasta, nao um arquivo.")
            html_path = htmls[0].resolve()
            if not args.output and not config.get("output"):
                output_path = html_path
            print(f"Encontrei um unico .html nela e vou usar: {html_path}")
        elif len(htmls) > 1:
            lista = "\n".join(f"  - {h.name}" for h in htmls)
            sys.exit(
                f"ERRO: '{html_path}' e uma pasta (nao um arquivo), e tem mais de\n"
                f"um .html dentro dela:\n{lista}\n\n"
                "Informe o caminho completo do arquivo .html certo, incluindo o\n"
                "nome do arquivo (ex: ...\\App RTA\\index.html)."
            )
        else:
            sys.exit(
                f"ERRO: '{html_path}' e uma pasta (nao um arquivo) e nao tem\n"
                "nenhum .html dentro dela. Informe o caminho completo do arquivo,\n"
                "incluindo o nome (ex: ...\\App RTA\\index.html)."
            )
    elif not html_path.exists():
        sys.exit(f"ERRO: HTML nao encontrado em: {html_path}")

    if usar_legado:
        xlsx_path = resolver_planilha(Path(xlsx_arg).resolve(), "Base Juntada")
        wb_unidade = abrir_planilha(xlsx_path, "Unidade+Dados")
        wb_dados = wb_unidade
    else:
        unidade_path = resolver_planilha(Path(unidade_arg).resolve(), "Unidade")
        dados_path = resolver_planilha(Path(dados_arg).resolve(), "Dados")
        wb_unidade = abrir_planilha(unidade_path, "Unidade")
        wb_dados = abrir_planilha(dados_path, "Dados")

    ws_unidade = escolher_aba(wb_unidade, UNIDADE_SHEET_CANDIDATOS, "Unidade")
    ws_dados = escolher_aba(wb_dados, DADOS_SHEET_CANDIDATOS, "Dados")

    unidade = extract_unidade(ws_unidade)
    dados = extract_dados(ws_dados)

    print(f"  -> {len(unidade)} RTAs (Unidade)")
    print(f"  -> {len(dados)} acoes (Dados)")

    unidade_json = json.dumps(unidade, ensure_ascii=False)
    dados_json = json.dumps(dados, ensure_ascii=False)

    tamanho_antes = html_path.stat().st_size
    mtime_antes = html_path.stat().st_mtime

    # Se o arquivo estiver marcado como "somente leitura" (comum em arquivos
    # sincronizados/baixados), tenta remover esse atributo automaticamente.
    try:
        os.chmod(html_path, stat.S_IWRITE | stat.S_IREAD)
    except Exception:
        pass  # se nao conseguir, o erro real aparecera abaixo de qualquer forma

    # Faz um backup do HTML antes de sobrescrever (facilita comparar/recuperar)
    backup_path = html_path.with_suffix(html_path.suffix + ".bak")
    try:
        conteudo_original = html_path.read_bytes()
    except PermissionError as e:
        sys.exit(
            f"ERRO: nao consegui LER o arquivo:\n  {html_path}\n\n"
            f"Mensagem do sistema: {e}\n\n"
            "Possiveis causas: o arquivo esta aberto em outro programa que o bloqueia,\n"
            "o antivirus esta escaneando/bloqueando o arquivo, ou falta permissao na pasta.\n"
            "Tente fechar programas que possam estar usando o arquivo e rodar novamente."
        )

    try:
        backup_path.write_bytes(conteudo_original)
        print(f"Backup criado em: {backup_path}")
    except PermissionError as e:
        sys.exit(
            f"ERRO: nao consegui GRAVAR em:\n  {backup_path}\n\n"
            f"Mensagem do sistema: {e}\n\n"
            "Possiveis causas:\n"
            "  - A pasta esta marcada como somente leitura, ou voce nao tem\n"
            "    permissao de escrita nela.\n"
            "  - O OneDrive esta sincronizando o arquivo/pasta neste momento\n"
            "    (espere o icone de nuvem terminar e tente de novo).\n"
            "  - O antivirus/Windows Defender esta bloqueando a escrita.\n"
            "  - O arquivo esta aberto em outro programa.\n\n"
            "Tente: clique com botao direito na pasta > Propriedades > desmarque\n"
            "'Somente leitura' (se disponivel), ou rode o VS Code como Administrador."
        )

    print(f"Atualizando HTML: {html_path}")
    try:
        update_html(html_path, unidade_json, dados_json, output_path)
    except PermissionError:
        sys.exit(
            f"ERRO: nao consegui gravar em {output_path}.\n"
            "Isso normalmente acontece se o arquivo estiver aberto em outro programa\n"
            "(editor de texto, Excel, etc.) ou se o OneDrive estiver bloqueando o arquivo.\n"
            "Feche o arquivo em outros programas e tente novamente."
        )
    except ValueError as e:
        sys.exit(f"ERRO: {e}")

    tamanho_depois = output_path.stat().st_size
    mtime_depois = output_path.stat().st_mtime

    print("-" * 60)
    print(f"Tamanho do arquivo antes:  {tamanho_antes} bytes")
    print(f"Tamanho do arquivo depois: {tamanho_depois} bytes")
    if output_path == html_path and mtime_depois == mtime_antes:
        print("AVISO: a data de modificacao do arquivo nao mudou. Algo pode ter dado errado.")
    print(f"Concluido! Arquivo atualizado em: {output_path}")
    print("\nSe o navegador ainda mostrar os dados antigos, feche a aba e abra o")
    print("arquivo de novo (ou aperte Ctrl+F5) -- o navegador pode ter guardado")
    print("uma versao em cache da pagina.")

    if not args.no_push:
        repo_dir = encontrar_repo_git(output_path)
        if repo_dir is None:
            print("-" * 60)
            print("AVISO: nao encontrei um repositorio git (.git) na pasta do HTML")
            print("nem em nenhuma pasta acima dela. Pulei o envio automatico para o GitHub.")
        else:
            git_sincronizar(repo_dir)
    else:
        print("(--no-push usado: pulei o envio automatico para o GitHub)")


if __name__ == "__main__":
    main()
